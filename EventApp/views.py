# inlcude the various features which are to be used in Views here
import copy
import datetime
import json
import re
import string
from io import BytesIO
from django.http import HttpRequest
import openpyxl
import qrcode
import sweetify
from django.conf import settings
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib.sites.shortcuts import get_current_site
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage
from django.db import IntegrityError
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from instamojo_wrapper import Instamojo

from EventApp.decorators import *
from GandharvaWeb19 import settings
from GandharvaWeb19.settings import BASE_DIR
from .email_sender import send_email
from .forms import *
from .token import *

from django.http.response import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from webpush import send_user_notification
import json


# @staff_user
def campaigning_excel(request):
    if request.method == "POST":
        all_transactions = Transaction.objects.filter(Q(status='Credit') | Q(status='Cash')).order_by('date')
        arg = {
            'transaction': all_transactions
        }
        if not request.POST.get('check'):
            return render(request, 'user/TableToExcel.html', arg)
        wb = openpyxl.Workbook()
        sheet = wb.active
        columns = ['Participant Name', 'Event', 'Phone No.', 'Date', 'Email Id', 'Mode', 'Refral Person', 'College']
        merge_col = ['Credit', 'Cash']
        heading_row_num = 1
        data_starting_number = 3
        merge_no = 5
        counter = 0
        sheet.column_dimensions['E'].width = 35
        for each_column in columns:
            if counter == merge_no:
                counter = counter + 1
                merge_no = merge_no + 1
                curr_cell = sheet.merge_cells(start_row=heading_row_num, start_column=merge_no, end_row=heading_row_num,
                                              end_column=merge_no + 1)
                sheet.cell(row=heading_row_num, column=merge_no).value = each_column
                # print(each_column)
                # print(merge_no)
                num = merge_no
                for col in merge_col:
                    sheet.cell(row=heading_row_num + 1, column=num).value = col
                    num = num + 1
                # print(counter, " ", each_column)
            else:
                curr_cell = sheet.cell(row=heading_row_num, column=counter + 1)
                curr_cell.value = each_column
                # print(each_column)
                # print(counter," ",each_column)
            counter = counter + 1

        flag = 1
        row = 0
        for each_transaction in all_transactions:
            # print(each_transaction.date)
            flag = 1
            if request.POST.get('date'):
                date = request.POST.get('date')
                # print(date)
                print(str(each_transaction.date), " ", str(date))
                if str(each_transaction.date) == str(date):
                    print(1)
                else:
                    flag = 0
            if flag == 1:
                if each_transaction.team.referral:
                    values = [each_transaction.team.user.first_name + " " + each_transaction.team.user.last_name,
                              each_transaction.receipt.event.event_name,
                              each_transaction.team.user.user_phone,
                              str(each_transaction.date),
                              each_transaction.team.user.email,
                              each_transaction.receipt.event.entry_fee,
                              each_transaction.team.referral.first_name + " " + each_transaction.team.referral.last_name,
                              each_transaction.team.user.user_coll.name]
                else:
                    values = [each_transaction.team.user.first_name + " " + each_transaction.team.user.last_name,
                              each_transaction.receipt.event.event_name,
                              each_transaction.team.user.user_phone,
                              str(each_transaction.date),
                              each_transaction.team.user.email,
                              each_transaction.receipt.event.entry_fee,
                              "",
                              each_transaction.team.user.user_coll.name]
                col = 0
                row = row + 1
                for each_value in values:
                    if col == (merge_no - 1):
                        if each_transaction.status == "Credit":
                            sheet.cell(row=row + data_starting_number, column=col + 1).value = each_value
                            # curr_cell.value = each_value
                        else:
                            sheet.cell(row=row + data_starting_number, column=col + 2).value = each_value
                            # curr_cell.value = each_value
                        col = col + 1
                        print(each_value)
                    else:
                        sheet.cell(row=row + data_starting_number, column=col + 1).value = each_value
                        # curr_cell.value = each_value
                    col = col + 1
        current_site = get_current_site(request)
        pathw = '/media/CampaignData.xlsx'
        wb.save(BASE_DIR + '/media/CampaignData.xlsx')
        if request.POST.get('check'):
            return (redirect(pathw))
        else:
            return render(request, 'user/TableToExcel.html', arg)

    else:
        all_transactions = Transaction.objects.filter(Q(status='Credit') | Q(status='Cash')).order_by('date')
        arg = {
            'transaction': all_transactions
        }
        return render(request, 'user/TableToExcel.html', arg)


def volunteer_excel(request):
    camp_team = RoleMaster.objects.get(name="Campaigning Team")
    all_camp = RoleAssignment.objects.filter(role=camp_team)
    wb = openpyxl.Workbook()
    sheet = wb.active
    columns = ['Volunteer Name', 'Phone No.']

    heading_row_num = 1

    data_starting_number = 3

    for counter, each_column in enumerate(columns):
        curr_cell = sheet.cell(row=heading_row_num, column=counter + 1)
        curr_cell.value = each_column

    for row, each_camp in enumerate(all_camp):
        values = [each_camp.user.first_name + " " + each_camp.user.last_name,
                  each_camp.user.user_phone]
        for col, each_value in enumerate(values):
            curr_cell = sheet.cell(row=row + data_starting_number, column=col + 1)
            curr_cell.value = each_value

    current_site = get_current_site(request)
    pathw = '/media/VolunteerData.xlsx'
    # return HttpResponse(BASE_DIR + '/media/CampaignData.xlsx')
    wb.save(BASE_DIR + '/media/VolunteerData.xlsx')
    arg = {
        'filename': pathw,
        'camp_teams': all_camp

    }
    # return HttpResponse()
    return render(request, 'user/Campaign_volunteer_excel.html', arg)


def offline(request):
    return render(request, 'gandharva/offline.html', {})


# Home page Functionality
def home(request):
    userget = 0
    if request.user and (not request.user.is_anonymous):
        user = request.user
        role = user.roleassignment_set.all()
        if len(role) > 1:
            return HttpResponse('Multiple Role')
        else:
            try:
                role = role[0]
                if role.role.name == "Jt Campaigning Head" or role.role.name == "Campaigning Head":
                    userget = 1
                if role.role.name == "Jt Event Head" or role.role.name == "Event Head":
                    userget = 2
                if role.role.name == "Jt Publicity Head" or role.role.name == "Publicity Head":
                    userget = 3
                if role.role.name == "Participant-Live":
                    userget = 4
            except:
                pass

    global_objects = EventDepartment.objects.filter(department=6)
    event_name = []
    for global_object in global_objects:
        event_name.append(global_object.event.event_name)

    args = {
        'events': Department.objects.all().order_by("rank"),
        'sponsors': SponsorMaster.objects.all().order_by('sponsor_rank'),
        'carouselImage': Carousel.objects.all(),
        'gandharvaDate': GandharvaHome.objects.get(title__startswith="Date").data,
        'About': GandharvaHome.objects.get(title__startswith="About").data,
        'role': userget,
        'global_events': event_name,
        'timelines': Event_days.objects.all().order_by("date")
    }
    sweetify.sweetalert(request, 'Westworld is awesome',
                        text='Really... if you have the chance - watch it! persistent = I agree!')
    return render(request, 'gandharva/index.html', args)


# ComingSoon Page
def coming_soon(request):
    arg = {
        'carouselImage': Carousel.objects.all(),
        'gandharvaDate': 'March 20, 2019'
    }
    return render(request, 'gandharva/comingSoon.html', arg)


# Events page of all Departments
def event_register(request):
    if request.GET:
        dept = request.GET.get('dept')
        dept_choose = Department.objects.get(name=dept)
        if "cultural" in dept_choose.name.lower() and not request.user.is_active:
            cultural = 0
        else:
            cultural = 1
        args1 = {
            'pageTitle': dept,
            'events': EventDepartment.objects.filter(department=dept_choose).order_by('event__rank'),
            'dept_choosen': dept_choose,
            'cultural': cultural
        }
        return render(request, 'events/newEvent.html', args1)


# Payment success
def success(request):
    if request.method == 'GET':
        # print("Enter success")
        payment_id = request.GET.get('payment_id')
        payment_status = request.GET.get('payment_status')
        payment_request_id = request.GET.get('payment_request_id')
        insta = InstamojoCredential.objects.latest('pk')
        api2 = Instamojo(api_key=insta.key,
                         auth_token=insta.token)
        response2 = api2.payment_request_payment_status(payment_request_id, payment_id)
        # print(response2)
        # print(response2['payment_request']['purpose'])  # Purpose of Payment Request
        # print(response2['payment_request']['payment']['status'])  # Payment status
        eid = request.GET.get("eid")
        event = EventMaster.objects.get(event_id=eid)
        user = MyUser.objects.get(email=response2['payment_request']['email'])
        try:
            transaction2 = Transaction.objects.get(transaction_id=payment_id)
        except(IntegrityError, ObjectDoesNotExist):
            transaction2 = None
            # transaction2=Transaction.objects.get(transaction_id=payment_id)
            # print(transaction2)

        if transaction2 is None:
            receipt = Receipt()
            team = Team()
            transaction = Transaction()
            receipt.name = response2['payment_request']['payment']['buyer_name']
            receipt.event = event
            receipt.save()
            team.receipt = receipt
            team.user = user
            if request.GET.get('ref') != "0":
                try:
                    referral = MyUser.objects.get(username=request.GET.get('ref'))
                    team.referral = referral
                except:
                    pass
            team.save()
            transaction.transaction_id = payment_id
            transaction.transaction_request_id = payment_request_id
            transaction.instrment_type = response2['payment_request']['payment']["instrument_type"],
            transaction.billing_instrument = response2['payment_request']['payment']["billing_instrument"]
            transaction.status = payment_status
            transaction.receipt = receipt
            transaction.date = datetime.date.today()
            transaction.time = datetime.datetime.now().time()
            transaction.team = team
            transaction.save()
            # Generate QR code if transaction is success full
            if transaction.status == "Credit":
                # print("credit ...")
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
                # content = "event:" + event.event_name + ", user:" + user.username

                # While loop for generating a unique refral code for user
                unique_event_code = ''
                while True:
                    try:
                        unique_event_code = ''.join(random.choice(string.ascii_uppercase) for _ in range(3)) + str(
                            random.randint(100, 999))
                        Team.objects.get(Refral_Code=unique_event_code)
                    except:
                        break
                content = {
                    'Event': event.event_name,
                    'Username': user.username,
                    'Name': receipt.name,
                    'Amount': receipt.event.entry_fee,
                    'Status': transaction.status,
                    'Mode': transaction.billing_instrument
                }
                qr.add_data(json.dumps(content))
                img = qr.make_image(fill_color="black", back_color="white")
                # img.save(user.username + event.event_name + "png")
                thumb_io = BytesIO()
                img.save(thumb_io, format='JPEG')
                team.QRcode.save('ticket-filename.jpg', File(thumb_io), save=False)

                team.Refral_Code = unique_event_code
                team.save()

                # Event Receipt Mail

                mail_subject = 'You have registered for ' + event.event_name + ' using cash payment'
                message = render_to_string('events/receiptCashPayment.html', {
                    'user': user,
                    'event': event,
                    'team': team,
                    'transaction': transaction,
                })

                send_email(user.email, mail_subject, message, [team.QRcode.path])

            if transaction.status == "Credit":
                return render(request, 'user/paymentSsuccess.html', {'user': user})
            elif transaction.status == "Failed":
                return render(request, 'user/paymentFailed.html', {'user_id': user.pk})
            teams = reversed(Team.objects.filter(user=user).reverse())
            # print(teams)
        else:
            return redirect('/')
    else:
        print("ERROR")


def hear_about_us(request):
    if request.method == 'GET':
        user_id = request.GET.get('user_id')
        source = request.GET.get('source')

        obj = HearAboutUs(user_id=user_id, source=source)
        obj.save()
        return redirect('home')


@staff_user
def all_participants(request):
    role = RoleAssignment.objects.get(user=request.user.id)
    if role.role.name == 'Event Head':
        eventid = role.event.event_id
        # print(eventid)
        # receipt = Receipt.objects.filter(event=eventid)
        # receipt.event.event_id = event_id
        receipts = Receipt.objects.filter(event=eventid)
        # print(receipts)

        """participants = []
        for receipt in receipts:
            user = receipt.transaction_set.all()
            participants.append(user)
        print("----------------------------")
        print(participants)
        for participant in participants:
            print(participant)
        """

        participants = Transaction.objects.all()
        arg = {
            'participants': participants,
        }
        return render(request, 'events/all_participants.html', arg)


def mail_participants(request):
    if request.method == 'GET':
        emails = request.GET.getlist('email')
        return render(request, 'events/mail_participants.html', {'emails': emails})

    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')

        # regular expression to convert string to list
        to_email = re.findall(r'[\w\.-]+@[\w\.-]+', request.POST.get('emails'))

        send_email(to_email, subject, message)
        return render(request, 'events/mail_response.html')


# Details of Individual Events
def details(request):
    if request.method == 'POST':
        insta = InstamojoCredential.objects.latest('pk')
        api = Instamojo(api_key=insta.key,
                        auth_token=insta.token, )
        event_id = request.POST.get('event_id')
        userEmail = request.POST.get('userEmail')
        event = EventMaster.objects.get(pk=event_id)
        user = MyUser.objects.get(email=userEmail)
        current_site = get_current_site(request)
        head = 'http://'
        if settings.USE_HTTPS:
            head = 'https://'
        response = api.payment_request_create(
            amount=event.entry_fee,
            purpose=event.event_name,
            send_email=False,
            send_sms=False,
            email=user.email,
            phone=user.user_phone,
            buyer_name=user.first_name + " " + user.last_name,

            redirect_url=head + current_site.domain + "success?eid=" + event_id
        )
        # print the long URL of the payment request.
        # print(response['payment_request']['longurl'])
        # print the unique ID(or payment request ID)
        # print(response['payment_request']['id'])
        # print(response['payment_request']['purpose'])
        # print(response['payment_request']['amount'])

        return redirect(response['payment_request']['longurl'])
    else:
        event_name = request.GET.get('event')
        dept_choose = EventDepartment.objects.get(
            event=EventMaster.objects.get(event_name__startswith=event_name)).department
        if "cultural" in dept_choose.name.lower() and not request.user.is_active:
            cultural = 0
        else:
            cultural = 1
        arg = {
            'events_list': EventMaster.objects.all().order_by('rank'),
            'pageTitle': EventMaster.objects.get(event_name__startswith=event_name).event_name,
            'event': EventMaster.objects.get(event_name__startswith=event_name),
            'dept': EventDepartment.objects.get(event=EventMaster.objects.get(event_name__startswith=event_name)),
            'cultural': cultural
        }
        return render(request, 'events/category1Event1.html', arg)


# ContactUs View (Form created)
def contactus(request):
    success_form = 2
    if request.method == 'POST':
        form = ContactUsForm(request.POST)
        if form.is_valid():
            msg = request.POST.get('user_message')
            name = request.POST.get('user_name')
            user_email = request.POST.get('user_id')
            category = request.POST.get('category')
            form.save()
            success_form = 1
            mail_subject = 'The person' + name + ' has contacted us'
            message = render_to_string('gandharva/contact-us-mail.html', {
                'user': name,
                'category': category,
                'id': user_email,
                'msg': msg,
            })
            send_email('gandharvaviitpune@gmail.com', mail_subject, message)

        else:
            # print(form.errors)
            success_form = 0
    else:
        form = ContactUsForm()

    return render(request, 'gandharva/contactus.html', {'form': form, 'success_form': success_form})


# Registration for normal User and log in user after registration Immediately
@user_passes_test(lambda u: u.is_superuser)
def register(request):
    dept = Department.objects.all()
    coll = College.objects.all()
    year = College_year.objects.all()
    if request.method == 'POST':
        form = UserRegistration(request.POST, request.FILES)

        new_email = request.POST.get('email')

        try:
            old_user = MyUser.objects.get(email=new_email)
        except:
            old_user = None

        if old_user is not None and old_user.is_active is False:
            old_user.delete()

        elif old_user is not None and old_user.is_active is True:
            args = {
                'error': "You have already registered and your email is verified too. Enter email to reset your password."
            }
            return render(request, "user/reset_password.html", args)

        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            password = form.cleaned_data.get('password')
            user.set_password(password)
            user.user_phone = form.cleaned_data.get('user_phone')
            user.save()
            current_site = get_current_site(request)
            token1 = account_activation_token.make_token(user)
            message = render_to_string('user/acc_active_email.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)).decode(),
                'token': token1,
            })
            user.token1 = str(token1)
            user.save()
            mail_subject = 'Activate your account to continue.'
            to_email = form.cleaned_data.get('email')
            send_email(to_email, mail_subject, message)
            return render(request, 'user/AccountConfirm.html')
            # else:
            # print(form.errors, "heere")
    else:
        form = UserRegistration()

    return render(request, 'events/register.html', {'form': form, 'colleges': coll, 'depts': dept, 'years': year})


# Activates the user after clicking on the email link
def activate(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = MyUser.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, user.DoesNotExist):
        user = None
    if user is not None:
        if user.token1 == token:
            user.is_active = True
            user.token1 = None
        user.save()
        # return redirect('home')
        return render(request, 'user/accountActivate.html')
    else:
        return HttpResponse('You have already confirmed your email id. Activation link is invalid!')


# logout Option View appears only after login
def user_logout(request):
    logout(request)
    return redirect('home')


# Login for user to Existing Account
def user_login(request):
    if request.user.is_active:
        return redirect('home')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            try:
                usercheck = MyUser.objects.get(username=username)
            except ObjectDoesNotExist:
                usercheck = None
            if usercheck is not None:
                if not usercheck.is_active:
                    # print("your account is inactive")
                    messages.error(request, 'Email not verified, please verify your email to login')
                    return render(request, 'events/login.html', {})

                else:
                    user = authenticate(username=username, password=password)
                    if user is not None:
                        if user.is_active:
                            # print("io")
                            login(request, user)
                            return redirect('home')
                        else:
                            messages.error(request, 'Invalid Username or Password')
                            return render(request, 'events/login.html', {})
                    else:
                        messages.error(request, 'Invalid Username or Password')
                        return render(request, 'events/login.html', {})

            else:
                messages.error(request, 'Invalid Username or Password')
                return render(request, 'events/login.html', {})
        else:
            return render(request, 'events/login.html', {})


@staff_user
def myaction(request):
    stat = 0
    role = RoleAssignment.objects.get(user=request.user)
    if request.user.is_staff:
        if role.role.name == "Event Head":
            stat = 1
            event_id = 0
        publicity = 0
        if role.role.name == "Publicity Head":
            publicity = 1
        if request.GET.get('event_id'):
            event_id = request.GET.get('event_id')
        args = {
            'button_name': '',
            'urlaccess': campaign,
            'roles': role.role,
            'stat': stat,
            'event_id': event_id,
            'publicity' : publicity,
        }
        return render(request, 'user/myactions.html', args)
    # if role.role.name == 'Event Head':
    #     args = {
    #         'button_name': 'All Participants',
    #         'roles': role.role
    #     }
    #     return render(request, 'user/myactions.html', args)
    # else:
    #     args = {
    #         'button_name': "No Actions",
    #         'urlaccess': None,
    #         'roles': role.role
    #     }
    # return render(request, 'user/myactions.html', args)


def payment(request):
    return render(request, 'user/paymentDetails.html', {})


# Head Login View only to be used for Heads
# @user_passes_test(lambda u: u.is_superuser)
def register_head(request):
    print("enter")
    Roles = RoleMaster.objects.all()
    role_categories = Role_category.objects.all()
    dept = Department.objects.all()
    coll = College.objects.all()
    year = College_year.objects.all()
    events = EventMaster.objects.all()
    if request.method == 'POST':
        id = request.POST.get('event')
        event = EventMaster.objects.get(pk=id)
        user_stat = 0
        # try:
        #     old_user = MyUser.objects.get(email=userform.email)
        #     print(old_user)
        # except:
        #     old_user = None
        # try:
        #     old_user2 = MyUser.objects.get(coll_email=userform.coll_email)
        # except:
        #     old_user2 = None
        #
        # print(old_user2)
        # print(old_user)
        if MyUser.objects.filter(email=request.POST.get('email')).count() == 0:
            old_user = None
            print(old_user)
        else:
            old_user = MyUser.objects.get(email=request.POST.get('email'))
            print(old_user)
        if MyUser.objects.filter(coll_email=request.POST.get('coll__email')).count() == 0:
            old_user2 = None
            print(old_user2)
        else:
            old_user2 = MyUser.objects.get(coll_email=request.POST.get('coll__email'))
            print(old_user2)
        if old_user is not None and old_user.is_active is False:
            user_stat = 1
        elif old_user2 is not None and old_user2.is_active is False and old_user is None:
            user_stat = 2
        elif (old_user is not None and old_user.is_active is True) or (
                old_user2 is not None and old_user2.is_active is True):
            user_stat = 0
            args = {
                'error': "You have already registered and your email is verified too. Enter email to reset your password."
            }
            return render(request, "user/reset_password.html", args)
        userform = UserRegistration(request.POST, request.FILES)
        roleform = RoleMasterForm(request.POST)
        if userform.is_valid():
            if user_stat == 1:
                user = old_user
            elif user_stat == 2:
                user = old_user2
            else:
                user = userform.save(commit=False)
            password = userform.cleaned_data.get('password')
            coll_email = userform.cleaned_data.get('coll_email')
            user_mail = userform.cleaned_data.get('email')
            user_name = userform.cleaned_data.get('username')
            user_year = userform.cleaned_data.get('user_year')
            user_coll = userform.cleaned_data.get('user_coll')
            user_mobile = userform.cleaned_data.get('user_phone')
            user.coll_email = coll_email
            flag = 0
            if "@viit" in coll_email:
                if "@viit.ac.in" not in coll_email:
                    flag = 1
            if "@viit" in user_mail:
                if "@viit.ac.in" not in user_mail:
                    flag = 1

            if flag == 1:
                userform = UserRegistration()
                roleform = RoleMasterForm
                selected_roles = RoleMaster.objects.all().order_by('name')
                # print(selected_roles)
                return render(request, 'events/RegisterHead.html',
                              {'userform': userform, 'roleform': roleform, 'roles': Roles, 'depts': dept,
                               'colleges': coll,
                               'years': year, 'categories': role_categories, 'selected_roles': selected_roles,
                               'events': events})

            user.email = user_mail
            user_year = College_year.objects.get(title=user_year)
            user_coll = College.objects.get(name=user_coll)
            user.user_coll = user_coll
            user.user_year = user_year
            user.username = user_name
            user.user_phone = user_mobile
            user.set_password(password)
            user.is_active = False
            user.full_name = user.first_name + " " + user.last_name
            user.save()
            role = request.POST.get('role')

            roleassign = RoleAssignment()
            roleassign.user = user
            roleassign.role = RoleMaster.objects.get(name=role)

            # if role == "Event Head":
            #     event.head = user
            # elif role == "Jt Event Head":
            #     event.jt_head = user

            event.save()
            roleassign.event = event

            current_site = get_current_site(request)
            token1 = account_activation_token.make_token(user)
            message = render_to_string('user/acc_active_email_register_head.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)).decode(),
                'token': token1,
            })
            user.token1 = str(token1)
            token2 = account_activation_token.make_token(user)
            message2 = render_to_string('user/acc_active_email_register_head.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)).decode(),
                'token': token2,
            })
            user.token2 = str(token2)
            user.save()
            roleassign.save()
            mail_subject = 'Activate your account to continue.'
            to_email_one = userform.cleaned_data.get('email')
            to_email_two = userform.cleaned_data.get('coll_email')
            send_email(to_email_one, mail_subject, message)
            send_email(to_email_two, mail_subject, message2)
            return render(request, 'user/AccountConfirm.html')

            #       group = Group.objects.get(name='groupname')
            #      user.groups.add(group)
            # login(request, user, backend='social_core.backends.google.GoogleOAuth2')
            # else:
            # print(userform.errors)
            # print(roleform.errors)
    else:
        userform = UserRegistration()
        roleform = RoleMasterForm
    selected_roles = RoleMaster.objects.all().order_by('name')
    # print(selected_roles)
    return render(request, 'events/RegisterHead.html',
                  {'userform': userform, 'roleform': roleform, 'roles': Roles, 'depts': dept, 'colleges': coll,
                   'years': year, 'categories': role_categories, 'selected_roles': selected_roles, 'events': events})


# @staff_user
def event_head(request):
    if request.method == "POST":
        participants_selected = request.POST.getlist('participants')
        print(participants_selected)
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        if len(request.FILES) == 1:
            myfile = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.path(filename)
            send_email(participants_selected, subject, message, [uploaded_file_url])
        else:
            send_email(participants_selected, subject, message)
        event_id = request.POST.get('event_id')
        participants1 = Team.objects.filter(receipt__event__pk=event_id).values_list('user', flat=True).distinct()
        participants = []
        for p in participants1:
            participants.append(MyUser.objects.get(pk=p))
        return render(request, 'events/event_head.html', {'participants': participants, 'event_id': event_id})
    else:
        event_id = request.GET.get('event_id')
        participants1 = Team.objects.filter(receipt__event__pk=event_id).values_list('user', flat=True).distinct()
        participants = []
        for p in participants1:
            participants.append(MyUser.objects.get(pk=p))
        return render(request, 'events/event_head.html', {'participants': participants, 'event_id': event_id})


# @staff_user
def publicity_head(request):
    if request.method == "POST":
        participants_selected = request.POST.getlist('participants')
        print("selected", participants_selected)
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        if len(request.FILES) == 1:
            myfile = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.path(filename)
            send_email(participants_selected, subject, message, [uploaded_file_url])
        else:
            send_email(participants_selected, subject, message)
        participants1 = Team.objects.values_list('user', flat=True).distinct().all()
        participants = []
        for p in participants1:
            participants.append(MyUser.objects.get(pk=p))
        return render(request, 'events/publicity_head.html', {'participants': participants})
    else:
        participants1 = Team.objects.values_list('user', flat=True).distinct().all()
        participants = []
        for p in participants1:
            participants.append(MyUser.objects.get(pk=p))
        return render(request, 'events/publicity_head.html', {'participants': participants})


def load_roles(request):
    category_id = request.GET.get('role_categories')
    roles = Category_assign.objects.filter(category_id=category_id)
    return render(request, 'events/category_roles.html', {'selected_roles': roles})


def activate_register_head(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = MyUser.objects.get(pk=uid)

    except(TypeError, ValueError, OverflowError, user.DoesNotExist):
        user = None
    if user is not None:
        if user.token1 == token:
            user.token1 = None
        if user.token2 == token:
            user.token2 = None
        if user.token1 is None and user.token2 is None:
            user.is_active = True
            user.is_staff = True
        user.save()
        # return redirect('home')
        return render(request, 'user/accountActivate.html')
    else:
        return HttpResponse('You have already confirmed your email id. Activation link is invalid!')


def participant_event_register(request):
    stat = 0
    if request.method == 'POST':
        useremail = request.POST.get('email')
        eventId = request.POST.get('event_id')
        event = EventMaster.objects.get(event_id=eventId)
        if event.can_register:
            if useremail != "":
                if event.event_name == 'Marathon & Zumba':
                    temp_email = useremail
                    domain = temp_email.split('@')[1]
                    if domain != 'viit.ac.in':
                        stat = 1
                    else:
                        stat = 0
                if stat == 0:
                    otp = random.randint(100000, 999999)
                    message = render_to_string('user/OTP.html', {
                        'otp': otp
                    })
                    mail_subject = 'OTP for email verification.'
                    request.session['otp'] = otp
                    send_email(useremail, mail_subject, message)

                return render(request, 'events/participantEventRegister.html',
                              {'email': useremail, 'event_id': eventId, 'btndisable': True, 'stat': stat})
            else:
                return render(request, 'events/participantEventRegister.html',
                              {'event_id': eventId, 'email': useremail})
        else:
            return HttpResponse("Sorry this event is not available for registration")
    if request.method == 'GET':
        event_id = request.GET.get('event_id')
        event = EventMaster.objects.get(event_id=event_id)
        if event.can_register:
            return render(request, 'events/participantEventRegister.html', {'event_id': event_id})
        else:
            return HttpResponse("Sorry this event is not available for registration")


def verifyOTP(request):
    if request.method == 'POST':
        userEmail = request.POST.get('useremail')
        eventId = request.POST.get('event_id')
        event = EventMaster.objects.get(pk=eventId)
        event_dept = EventDepartment.objects.get(event=event).department.name
        if "cultural" in event_dept.lower():
            cultural = 1
        else:
            cultural = 0
        otpEntered = request.POST.get('otp')
        originalotp = str(request.session.get('otp'))
        # print(originalotp)
        # print("original otp", originalotp)
        if originalotp != otpEntered or len(originalotp) < 6:
            error = "Invalid OTP"
            return render(request, 'events/participantEventRegister.html',
                          {'error': error, 'event_id': eventId, 'email': userEmail})

        else:
            request.session['otp'] = ""
            coll = College.objects.all().order_by('name')
            if MyUser.objects.filter(email=userEmail).count() == 0:
                if MyUser.objects.filter(coll_email=userEmail).count() == 0:
                    ifuser = None

                else:
                    ifuser = MyUser.objects.get(coll_email=userEmail)
            else:
                ifuser = MyUser.objects.get(email=userEmail)
            readm = "readonly"
            return render(request, 'events/participantDetails.html',
                          {'event': event, 'colleges': coll, 'email_participant': userEmail, 'present_user': ifuser,
                           'readm': readm, 'cultural': cultural})


def participant_details(request):
    if request.method == "POST":

        participant_email = request.POST.get('email')
        event_id = request.POST.get('event_id')
        # print(event_id)
        # print("POst mail:", participant_email)
        form = PaymentForm(request.POST)
        coll = College.objects.all().order_by('name')

        if MyUser.objects.filter(email=participant_email).count() == 0:
            if MyUser.objects.filter(coll_email=participant_email).count() == 0:
                ifuser = None

            else:
                ifuser = MyUser.objects.get(coll_email=participant_email)
        else:
            ifuser = MyUser.objects.get(email=participant_email)

        event_new = EventMaster.objects.get(pk=event_id)
        if len(request.POST.get('user_phone')) < 10:
            error = "Invalid mobile number"
            return render(request, 'events/participantDetails.html',
                          {'event': event_new, 'colleges': coll, 'email_participant': participant_email,
                           'present_user': ifuser, 'error': error})
        if request.POST.get('button_state') == "on":
            pass
        else:
            error = "You need to accept"
            return render(request, 'events/participantDetails.html',
                          {'event': event_new, 'colleges': coll, 'email_participant': participant_email,
                           'present_user': ifuser, 'error': error})
        if form.is_valid():
            if ifuser is None:
                user = form.save(commit=False)
                password = None
                user.set_password = password
                user.username = participant_email
                user.is_active = False
                user.save()

            # print("name", event_new.event_name)
            if MyUser.objects.filter(email=participant_email).count():
                user = MyUser.objects.get(email=participant_email)
            elif MyUser.objects.filter(coll_email=participant_email).count():
                user = MyUser.objects.get(coll_email=participant_email)
            if request.POST.get('card'):
                # print("towards Payment")
                current_site = get_current_site(request)
                # print("here", current_site.domain)
                refer = request.POST.get('refer')
                try:
                    referral = MyUser.objects.get(username=refer)
                except(IntegrityError, ObjectDoesNotExist):
                    refer = "0"

                insta = InstamojoCredential.objects.latest('pk')
                # print(insta.key, insta.token, insta.salt)
                api = Instamojo(api_key=insta.key,
                                auth_token=insta.token, )
                head = 'http://'
                if settings.USE_HTTPS:
                    head = 'https://'

                response = api.payment_request_create(
                    amount=event_new.entry_fee,
                    purpose=event_new.event_name,
                    send_email=False,
                    send_sms=False,
                    email=user.email,
                    phone=user.user_phone,
                    buyer_name=user.first_name + " " + user.last_name,
                    # redirect_url="http://127.0.0.1:8000/success?eid=" + event_id + "&ref=" + str(refer)
                    redirect_url=head + current_site.domain + "/" + "success?eid=" + event_id + "&ref=" + str(
                        refer) + "success?eid=" + event_id + "&ref=" + str(refer)

                )
                # print the long URL of the payment request.
                # print(response)
                # print(response['payment_request']['longurl'])
                # print the unique ID(or payment request ID)
                # print(response['payment_request']['id'])
                # print(response['payment_request']['purpose'])
                # print(response['payment_request']['amount'])
                return redirect(response['payment_request']['longurl'])
            elif request.POST.get('cash'):
                cashpayment(request, event_new, user)
                teams = reversed(Team.objects.filter(user=user).reverse())
                # print("CAAAAA")
                return render(request, 'user/registeredEvents.html', {'teams': teams})
        else:
            # print(form.errors)
            error = form.errors
            return render(request, 'events/participantDetails.html',
                          {'event': event_new, 'colleges': coll, 'email_participant': participant_email,
                           'present_user': ifuser, 'error': error, 'form': form})
    else:
        return redirect('/')


@staff_user
def cashpayment(request, event_new, user):
    id = ""
    flag = 0
    while flag == 0:
        try:
            id = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(20))
            transaction = Transaction.objects.get(transaction_id=id)
        except(IntegrityError, ObjectDoesNotExist):
            flag = 1
    receipt = Receipt()
    team = Team()
    transaction = Transaction()
    receipt.name = user.first_name + " " + user.last_name
    receipt.event = event_new
    receipt.save()
    team.receipt = receipt
    team.user = user
    team.referral = request.user
    team.save()
    transaction.transaction_id = id
    transaction.transaction_request_id = id
    transaction.instrment_type = "Cash"
    transaction.billing_instrument = "Cash"
    transaction.status = "Cash"
    transaction.receipt = receipt
    transaction.date = datetime.date.today()
    transaction.time = datetime.datetime.now().time()
    transaction.team = team
    transaction.save()
    # Generate QR code if transaction is success full
    if transaction.status == "Cash":
        # print("cashhh")
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
        content = "event:" + event_new.event_name + ", user:" + user.username
        content = {
            'Event': event_new.event_name,
            'Username': user.username,
            'Name': receipt.name,
            'Amount': receipt.event.entry_fee,
            'Status': transaction.status,
            'Mode': transaction.billing_instrument
        }
        qr.add_data(json.dumps(content))
        # qr.add_data(content)
        img = qr.make_image(fill_color="black", back_color="white")
        # img.save(user.username + event.event_name + "png")
        thumb_io = BytesIO()
        img.save(thumb_io, format='JPEG')
        team.QRcode.save('ticket-filename.jpg', File(thumb_io), save=False)

        # While loop for generating a unique refral code for user
        while True:
            try:
                Refral_Code = ''.join(random.choice(string.ascii_uppercase) for _ in range(3)) + str(
                    random.randint(100, 999))
                Team.objects.get(Refral_Code=Refral_Code)
            except:
                break

        team.Refral_Code = Refral_Code

        team.save()

        mail_subject = 'You have registered for ' + event_new.event_name + ' using cash payment'
        message = render_to_string('events/receiptCashPayment.html', {
            'user': user,
            'event': event_new,
            'team': team,
            'transaction': transaction,
        })
        send_email(user.email, mail_subject, message, [team.QRcode.path])


@staff_user
def Profile(request):
    user = request.user
    if request.method == 'POST':
        if request.method == 'POST' and len(request.FILES) == 1:
            prof_img = request.FILES['prof_img']
            # print(request.FILES['prof_img'])
            user.prof_img = prof_img
        user_phone = request.POST.get('user_phone')
        user.user_phone = user_phone
        user.save()
    return render(request, 'user/userProfile.html')


@staff_user
def registered_events(request):
    teams = Team.objects.filter(user=request.user)
    return render(request, 'user/registeredEvents.html', {'teams': teams})


def Payment_Details(request):
    return render(request, 'user/paymentDetails.html')


# def cashPayment(request):
#     coll = College.objects.all()
#     year = College_year.objects.all()
#     event_id = request.GET.get('event_id')
#     campaign_name = request.GET.get('userEmail')
#     if request.method == 'POST':
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             user = form.save(commit=False)
#             firstname = form.cleaned_data.get('first_name')
#             username = form.cleaned_data.get('username')
#             password = None
#             user.set_password=password
#             user.is_active=False
#             user.save()
#             receipt = Receipt()
#             receipt.event = EventMaster.objects.get(pk=event_id)
#             receipt.name = firstname
#             team = Team()
#             team.receipt=receipt
#             team.user=user
#             team.referral= campaign_name
#             receipt.save()
#             team.save()
#         else:
#             print(form.errors)
#
#     else:
#         form = PaymentForm()
#         event = EventMaster.objects.get(pk=event_id)
#     return render(request, 'events/cashPayment.html',{'form':form,'event':event,'colleges':coll,'years':year})

def team_details(request):
    event = request.GET.get('event')
    event_choose = EventMaster.objects.get(event_name=event)
    if request.method == 'GET':
        form = TeamDetailsForm()
    if request.method == 'POST':
        form = TeamDetailsForm(request.POST)
        if form.is_valid():
            form.save()
            # else:
            # print(form.errors)

    return render(request, 'events/TeamDetails.html', {'form': form, 'event': event_choose})


def reset_password(request):
    if request.method == 'POST':
        email_to_reset = request.POST.get('email')
        try:
            user = MyUser.objects.get(email=email_to_reset)
        except:
            return HttpResponse("Enter your correct email.Go back to enter the email.")

        current_site = get_current_site(request)
        token2 = account_activation_token.make_token(user)
        message = render_to_string('user/reset_password_email.html', {
            'user': user,
            'domain': current_site.domain,
            'uid': urlsafe_base64_encode(force_bytes(user.pk)).decode(),
            'token': token2,
        })
        user.token2 = str(token2)
        user.save()

        mail_subject = 'Reset Password'
        email = EmailMessage(mail_subject, message, to=[user.email])
        send_email(user.email, mail_subject, message, otp=1)
        return HttpResponse("Mail has been send. Click on the email link to reset password")
    else:
        return render(request, "user/reset_password.html")


def reset_password_new(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = MyUser.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, user.DoesNotExist, IntegrityError, ObjectDoesNotExist):
        user = None
    if user is not None:
        # return HttpResponse(user.token2 + 'a<br>' + token + 'b<br>')
        if str(user.token2) == str(token):
            args = {
                'user_new': user,
            }
            return render(request, 'user/new_password.html', args)
        else:
            return render(request, "user/reset_password.html")
    return HttpResponse("You have already reset your password")


def new_password(request):
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_new_password = request.POST.get('confirm_new_password')
        id = request.POST.get('pk')
        user = MyUser.objects.get(pk=id)
        if new_password == confirm_new_password:
            user.set_password(new_password)
            user.token2 = None
            user.save()
            return render(request, 'events/login.html', {})

    return render(request, 'user/new_password.html')


@staff_user
def change_password(request):
    status = True
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        user = authenticate(username=request.user.username, password=old_password)
        if user is not None:
            status = True
            return render(request, 'user/new_password.html', {'status': status})
        else:
            status = False
            messages.error(request, "Invalid Password", {'status': status})

    return render(request, 'user/change-password.html', {'status': status})


def send_username(request):
    status = 0
    if request.method == 'POST':
        user_email = request.POST.get('user-email')
        count = MyUser.objects.filter(email=user_email).count()
        if count:
            user = MyUser.objects.get(email=user_email)
            if user.is_active:
                message = render_to_string('user/username-email-sent.html',
                                           {
                                               'user': user,
                                           })
                mail_subject = 'Activate your account to continue.'
                to_email = user_email
                send_email(to_email, mail_subject, message)
                status = 1
                return render(request, 'user/send-username.html', {'status': status})
            else:
                status = 3
                return render(request, 'user/send-username.html', {'status': status})
        else:
            status = -1
            return render(request, 'user/send-username.html', {'status': status})
    return render(request, 'user/send-username.html', {'status': status})


# Important Notes:
# to get user role from models
# userget = RoleAssignment.objects.get(user=request.user.id)
#   print (userget.role)

def other_uploads(request):
    juniors = AssignSub.objects.filter(rootuser=request.user).count()
    if juniors:
        juniors = AssignSub.objects.filter(rootuser=request.user).order_by('subuser')
    else:
        juniors = None
    return render(request, 'user/other-uploads.html', {'juniors': juniors})


def uploaded_docs(request):
    if request.POST:
        us = request.POST.get('junior')
        junior = MyUser.objects.get(username=us)
        doc_lists = fileDocument.objects.filter(user=junior).count()
        if doc_lists:
            doc_lists = fileDocument.objects.filter(user=junior).order_by("uploaded_at").reverse()
        else:
            doc_lists = None
        return render(request, 'user/uploaded-docs.html', {'docs': doc_lists, 'junior': junior})
    else:
        return render(request, 'user/uploaded-docs.html', {})


# Volunteer College Date Entry by Campaign Head
@user_Campaign_head
def AddVolunteer(request):
    campaign_team = RoleMaster.objects.get(name='Campaigning Team')
    if request.method == "POST":
        uid = request.POST.get('volunteers')
        cid = request.POST.get('colleges')
        user = MyUser.objects.get(pk=uid)
        college = College.objects.get(pk=cid)
        volunteer = Volunteer()
        volunteer.user = user
        volunteer.college = college
        volunteer.date = request.POST.get('campaignDate')
        volunteer.save()
        volunteerData = Volunteer.objects.all()
        volunteers = []
        roles = RoleAssignment.objects.all()
        for role in roles:
            if role.role == campaign_team:
                volunteers.append(role.user)
        colleges = College.objects.all()
        args = {
            'volunteers': volunteers,
            'colleges': colleges,
            'volunteerData': volunteerData
        }
        return render(request, 'events/campaignVolunteer.html', args)


class CategoryWise:
    category = None
    sponsors = []
    partners = []


def ourSponsors(request):
    dict = {}
    Categories = SponsorCategory.objects.all().order_by('category_rank')
    for c in Categories:
        all_sponsor = SponsorMaster.objects.filter(sponsor_category=c).order_by('sponsor_rank')
        dict[c] = all_sponsor
        c.sponsor_category
    current_site = get_current_site(request)
    print(current_site)
    current_site = str(current_site) + "/media/"
    args = {
        'dict': dict,
        'site': current_site
    }
    return render(request, 'gandharva/ourSponsors.html', args)

    # Accessing category wise data:
    # run loop for c in category
    # c.category.sponsor_category = to access category name
    # c.category.category_rank = to access to category_rank
    # c.sponsors = sponsors of current category and c.sponsors[index].sponsors_name to access sponsor_name
    # similarly for partners of each category


def our_team(request):
    obj = OurTeam.objects.all().count()
    if obj:
        obj = OurTeam.objects.all().order_by('rank')
    else:
        obj = None
    return render(request, 'gandharva/ourTeam.html', {'objs': obj})


# upload file view
@staff_user
def files(request):
    try:
        glbdoc = Document.objects.get(category=Document_type.objects.get(type="Global"))
    except(ObjectDoesNotExist):
        glbdoc = None
    current_doc = fileDocument.objects.filter(user=request.user).order_by("uploaded_at").reverse().count()
    if current_doc:
        current_doc = fileDocument.objects.filter(user=request.user).order_by("uploaded_at").reverse()
    else:
        current_doc = None
    # dictonary = {}
    # juniors = AssignSub.objects.filter(rootuser=request.user)
    # for junior in juniors:
    #     doc_list = fileDocument.objects.filter(user=junior.subuser).order_by("uploaded_at").reverse()
    #     dictonary[junior.subuser] = doc_list

    if request.method == 'POST':
        form = fileForm(request.POST, request.FILES)
        if request.method == 'POST' and len(request.FILES) == 1:
            if form.is_valid():
                f = form.save(commit=False)
                f.user = request.user
                f.fname = request.FILES['document'].name
                # print(request.FILES['document'].name)
                f.save()
            return render(request, 'events/fileExplorer.html', {
                'form': fileForm,
                'documents': current_doc,
                'global': glbdoc
            })
        else:
            return render(request, 'events/fileExplorer.html', {
                'form': fileForm,
                'documents': current_doc,
                'global': glbdoc
            })
    else:
        form = fileForm()
    return render(request, 'events/fileExplorer.html', {
        'form': form,
        'documents': current_doc,
        'global': glbdoc
    })


# Campaign Head Method
@staff_user
def campaign(request):
    if request.method == "POST":
        check = request.POST.get('criteria')
        # volunteer wise
        if check == "0":
            volunteers = []
            transactions = Transaction.objects.all()
            for transaction in transactions:
                try:
                    ref = transaction.team.referral.username
                    name = MyUser.objects.get(username=ref)
                except Exception:
                    continue
                c = 0
                for i in range(len(volunteers)):
                    if volunteers[i].username == ref:
                        c = 1
                        volunteers[i].count = volunteers[i].count + 1
                        volunteers[i].money = volunteers[i].money + transaction.receipt.event.entry_fee
                        break
                if c == 0:
                    v = Volunteerwise()
                    v.username = ref
                    v.name = name
                    v.count = 1
                    v.money = transaction.receipt.event.entry_fee
                    volunteers.append(v)
            total = 0
            total_amount = 0
            for c in volunteers:
                total = total + c.count
                total_amount = total_amount + c.money
            args = {
                'volunteers': volunteers,
                'total': total,
                'total_amount': total_amount
            }
            return render(request, 'events/campaigningData.html', args)
        elif check == "1":
            events = []
            domains = []
            names = []
            # numbers = []
            transactions = Transaction.objects.all()
            for transaction in transactions:
                if transaction.status == "Credit" or transaction.status == "Cash":
                    event = transaction.receipt.event
                    c = 0
                    for i in range(len(events)):
                        if events[i].event_id == event.event_id:
                            c = 1
                            events[i].count = events[i].count + 1
                            break
                    if c == 0:
                        e = Eventwise()
                        d = EventDepartment.objects.get(event=event)
                        e.dept_id = d.department.dep_id
                        e.domain_name = d.department.name
                        e.event_id = event.event_id
                        e.event_name = event.event_name
                        e.count = 1
                        events.append(e)
                        names.append(e.event_name)

            eve = EventMaster.objects.all()
            # print(names)
            for e in eve:
                if e.event_name not in names:
                    print(e.event_name)
                    new = Eventwise()
                    d = EventDepartment.objects.get(event=e)
                    new.dept_id = d.department.dep_id
                    new.domain_name = d.department.name
                    new.event_id = d.event.event_id
                    new.event_name = d.event.event_name
                    new.count = 0
                    events.append(new)
                    names.append(new.event_name)
            glob = 0
            events.sort(key=lambda x: x.count, reverse=True)
            for e in events:
                do = 0
                for d in domains:
                    if e.dept_id > 5:
                        if glob == 0:
                            dom = Domainwise()
                            dom.name = "Global"
                            dom.count = e.count
                            ev = []
                            ev.append(copy.deepcopy(e))
                            dom.events = ev
                            domains.append(copy.deepcopy(dom))
                            del dom
                            glob = 1
                            do = 1
                            break
                        elif glob == 1:
                            if d.name == "Global":
                                d.count = d.count + e.count
                                d.events.append(copy.deepcopy(e))
                                break
                            do = 1

                    if e.domain_name == d.name:
                        d.count = d.count + e.count
                        d.events.append(copy.deepcopy(e))
                        do = 1
                        break
                if do == 0:
                    dom = Domainwise()
                    dom.name = e.domain_name
                    dom.count = e.count
                    ev = []
                    ev.append(copy.deepcopy(e))
                    dom.events = ev
                    domains.append(copy.deepcopy(dom))
                    del dom
            total = 0
            for c in events:
                total = total + c.count
            domains.sort(key=lambda x: x.count, reverse=True)
            args = {
                'domains': domains,
                'events': events,
                'total': total
            }
            # print("")
            # print(events)

            role = RoleAssignment.objects.get(user=request.user.id)
            campaign_object = RoleMaster.objects.filter(name="Campaigning Head")[0]
            head_level = campaign_object.level
            if head_level <= role.role.level:
                return render(request, 'events/campaigningData.html', args)
            else:
                return HttpResponse('You are not allowed')
        elif check == "2":
            colleges = []
            transactions = Transaction.objects.all()
            for transaction in transactions:
                if transaction.status == "Credit" or transaction.status == "Cash":
                    college = transaction.team.user.user_coll
                    c = 0
                    for i in range(len(colleges)):
                        if colleges[i].name == college.name:
                            c = 1
                            colleges[i].count = colleges[i].count + 1
                            break
                    if c == 0:
                        c = Collegewise()
                        c.name = college.name
                        c.count = 1
                        colleges.append(c)
            total = 0
            for c in colleges:
                total = total + c.count
            args = {
                'colleges': colleges,
                'total': total
            }
            # print("")
            # print(colleges)
            return render(request, 'events/campaigningData.html', args)

        elif check == "3":
            volunteers = []
            roles = RoleAssignment.objects.all()
            for role in roles:
                if role.role.name == "Campaigning Team":
                    volunteers.append(role.user)
                    # print(role.user)
            colleges = College.objects.all()
            volunteerData = Volunteer.objects.all()
            args = {
                'volunteers': volunteers,
                'colleges': colleges,
                'volunteerData': volunteerData
            }
            return render(request, 'events/campaignVolunteer.html', args)

    else:
        role = RoleAssignment.objects.get(user=request.user.id)
        campaign_object = RoleMaster.objects.filter(name="Campaigning Head")[0]
        head_level = campaign_object.level
        return render(request, 'events/campaignHead.html', {'role': role, 'head_level': head_level})


def terms(request):
    terms = TermsConditons.objects.all()

    return render(request, 'gandharva/terms-and-conditions.html', {'terms': terms})


def policy(request):
    policy = TermsConditons.objects.all()

    return render(request, 'gandharva/privacy-policy.html', {'policys': policy})


class Volunteerwise:
    username = ""
    name = ""
    count = 0
    money = 0


class Eventwise:
    event_id = 0
    event_name = ""
    count = 0
    dept_id = 0
    is_change = 0


class Collegewise:
    cid = ""
    name = ""
    count = 0


def run_custom():
    users = MyUser.objects.all()

    except_list = ['Ajinkya', 'Anand', 'rrs', 'shivamdeshpande']
    ajinkya_user = MyUser.objects.get(username='Ajinkya')
    for each in users:
        if each.username in except_list:
            pass
        else:
            AssignSub.objects.create(rootuser=ajinkya_user,
                                     subuser=each)


@user_Campaign_head
def participant_live(request):
    events = []
    domains = []
    names = []
    numbers = []
    transactions = Transaction.objects.all()
    for transaction in transactions:
        if transaction.status == "Credit" or transaction.status == "Cash":
            event = transaction.receipt.event
            c = 0
            for i in range(len(events)):
                if events[i].event_id == event.event_id:
                    c = 1
                    events[i].count = events[i].count + 1
                    break
            if c == 0:
                e = Eventwise()
                d = EventDepartment.objects.get(event=event)
                e.dept_id = d.department.dep_id
                e.domain_name = d.department.name
                e.event_id = event.event_id
                e.event_name = event.event_name
                e.count = 1
                events.append(e)
                names.append(e.event_name)

    eve = EventMaster.objects.all()
    # print(names)
    for e in eve:
        if e.event_name not in names:
            print(e.event_name)
            new = Eventwise()
            d = EventDepartment.objects.get(event=e)
            new.dept_id = d.department.dep_id
            new.domain_name = d.department.name
            new.event_id = d.event.event_id
            new.event_name = d.event.event_name
            new.count = 0
            events.append(new)
            names.append(new.event_name)
    glob = 0
    for e in events:
        do = 0
        for d in domains:
            if e.dept_id > 5:
                if glob == 0:
                    dom = Domainwise()
                    dom.name = "Global"
                    dom.count = e.count
                    ev = []
                    ev.append(copy.deepcopy(e))
                    dom.events = ev
                    domains.append(copy.deepcopy(dom))
                    del dom
                    glob = 1
                    do = 1
                    break
                elif glob == 1:
                    if d.name == "Global":
                        d.count = d.count + e.count
                        d.events.append(copy.deepcopy(e))
                        break
                    do = 1

            if e.domain_name == d.name:
                d.count = d.count + e.count
                d.events.append(copy.deepcopy(e))
                do = 1
                break
        if do == 0:
            dom = Domainwise()
            dom.name = e.domain_name
            dom.count = e.count
            ev = []
            ev.append(copy.deepcopy(e))
            dom.events = ev
            domains.append(copy.deepcopy(dom))
            del dom
    total = 0
    for d in domains:
        total = total + d.count
        numbers.append(d.count)
        for e in d.events:
            numbers.append(e.count)
    request.session['numbers'] = numbers
    request.session['total'] = total

    args = {
        'domains': domains,
        'total': total,

    }
    return render(request, 'events/participant-live.html', args)


class Domainwise:
    name = ""
    events = []
    count = 0


def live(request):
    if request.method == "POST":
        is_khamosh = 0
        events = []
        domains = []
        names = []
        numbers = []
        numbersold = request.session.get('numbers')
        transactions = Transaction.objects.all()
        for transaction in transactions:
            if transaction.status == "Credit" or transaction.status == "Cash":
                event = transaction.receipt.event
                c = 0
                for i in range(len(events)):
                    if events[i].event_id == event.event_id:
                        c = 1
                        events[i].count = events[i].count + 1
                        break
                if c == 0:
                    e = Eventwise()
                    d = EventDepartment.objects.get(event=event)
                    e.dept_id = d.department.dep_id
                    e.domain_name = d.department.name
                    e.event_id = event.event_id
                    e.event_name = event.event_name
                    e.count = 1
                    events.append(e)
                    names.append(e.event_name)

        eve = EventMaster.objects.all()
        # print(names)
        for e in eve:
            if e.event_name not in names:
                print(e.event_name)
                new = Eventwise()
                d = EventDepartment.objects.get(event=e)
                new.dept_id = d.department.dep_id
                new.domain_name = d.department.name
                new.event_id = d.event.event_id
                new.event_name = d.event.event_name
                new.count = 0
                events.append(new)
                names.append(new.event_name)
        glob = 0
        for e in events:
            do = 0
            for d in domains:
                if e.dept_id > 5:
                    if glob == 0:
                        dom = Domainwise()
                        dom.name = "Global"
                        dom.count = e.count
                        ev = []
                        ev.append(copy.deepcopy(e))
                        dom.events = ev
                        domains.append(copy.deepcopy(dom))
                        del dom
                        glob = 1
                        do = 1
                        break
                    elif glob == 1:
                        if d.name == "Global":
                            d.count = d.count + e.count
                            d.events.append(copy.deepcopy(e))
                            break
                        do = 1

                if e.domain_name == d.name:
                    d.count = d.count + e.count
                    d.events.append(copy.deepcopy(e))
                    do = 1
                    break
            if do == 0:
                dom = Domainwise()
                dom.name = e.domain_name
                dom.count = e.count
                ev = []
                ev.append(copy.deepcopy(e))
                dom.events = ev
                domains.append(copy.deepcopy(dom))
                del dom

        total = 0
        i = 0
        for d in domains:
            total = total + d.count
            i = i + 1
            numbers.append(d.count)
            for e in d.events:
                if numbersold:
                    if numbersold[i] != e.count:
                        e.is_change = 1
                        print(e.event_name)
                    i = i + 1
                numbers.append(e.count)
        request.session['numbers'] = numbers
        if request.session.get('total'):
            sub = total - request.session.get('total')
            if sub == 10:
                is_khamosh = 1
                request.session['total'] = total
        args = {
            'domains': domains,
            'total': total,
            'khamosh': is_khamosh

        }
        return render(request, 'events/live.html', args)


def pariwartan(request):
    if request.method == 'POST':
        useremail = request.POST.get('email')
        if useremail != "":
            otp = random.randint(100000, 999999)
            message = render_to_string('user/OTP.html', {
                'otp': otp
            })
            mail_subject = 'OTP for email verification.'
            request.session['otp'] = otp
            send_email(useremail, mail_subject, message)

            return render(request, 'events/pariwartan.html',
                          {'email': useremail, 'btndisable': True})
        else:
            return render(request, 'events/pariwartan.html', {'email': useremail})
    if request.method == 'GET':
        return render(request, 'events/pariwartan.html', {})


def verifyOTP_event(request):
    vishwa = EventMaster.objects.get(event_name="VishwaParivartan")
    stats = 0
    if request.method == 'POST':
        userEmail = request.POST.get('useremail')
        otpEntered = request.POST.get('otp')
        originalotp = str(request.session.get('otp'))
        # print(originalotp)
        # print("original otp", originalotp)
        if originalotp != otpEntered or len(originalotp) < 6:
            error = "Invalid OTP"
            return render(request, 'events/pariwartan.html',
                          {'error': error, 'email': userEmail})

        else:
            request.session['otp'] = ""
            if MyUser.objects.filter(email=userEmail).count() == 0 or MyUser.objects.filter(
                    coll_email=userEmail).count() == 0:
                ifuser = None
                error = "Not participated in this Event!!"
                stats = 0
                participant = None
            else:
                error = None
                ifuser = MyUser.objects.get(email=userEmail)
                if Team.objects.filter(user=ifuser).count():
                    teams = Team.objects.filter(user=ifuser)
                    for team in teams:
                        if team.receipt.event == vishwa:
                            participant = team
                            stats = 1
                        else:
                            stats = 0
                else:
                    stats = 0
                    participant = None

            readm = "readonly"
            return render(request, 'events/pariwartan_upload.html',
                          {'error': error, 'present_user': ifuser,
                           'readm': readm, 'participant': participant, 'stats': stats})


def pariwartan_upload(request):
    vishwa = EventMaster.objects.get(event_name="VishwaParivartan")
    usermail = request.POST.get('user')
    user = MyUser.objects.get(email=usermail)
    if Team.objects.filter(user=user).count():
        teams = Team.objects.filter(user=user)
        for team in teams:
            if Transaction.objects.get(team=team).receipt.event == vishwa:
                participant = team
    if request.method == 'POST' and len(request.FILES) == 1:
        usermail = request.POST.get('user')
        user = MyUser.objects.get(email=usermail)
        doc = request.FILES['doc']
        if Pariwartan.objects.filter(user=user).count():
            pariwartan = Pariwartan.objects.get(user=user)
            pariwartan.doc = doc
            pariwartan.save()
            stats = 2
        else:
            pariwartan = Pariwartan()
            pariwartan.user = user
            pariwartan.doc = doc
            pariwartan.save()
            stats = 2
    elif request.method == 'POST':
        stats = 1
    return render(request, 'events/pariwartan_upload.html', {'stats': stats, 'participant': participant})

@user_passes_test(lambda u: u.is_superuser)
def qr_verify(request):
    stat = 5
    selected = None
    if request.method == 'POST':
        if "qrcode" in request.POST:
            body = request.POST.get('textqr')
            try:
                data = json.loads(body)
                try:
                    event = data['event']
                    name = data['username']
                except:
                    event = data['Event']
                    name = data['Username']
                otpEntered = request.POST.get('textqr')
                print(otpEntered)
                print(event)
                print(name)
            except:
                stat = 2
            if stat is not 2:
                user = MyUser.objects.get(username=name)
                eventname = EventMaster.objects.get(event_name=event)
                if Team.objects.filter(user=user).count():
                    teams = Team.objects.filter(user=user)
                    for team in teams:
                        if team.receipt.event == eventname:
                            if team.ispresent is False:
                                selected = team
                                team.ispresent = True
                                team.save()
                                stat = 1
                            else:
                                stat = 3
                        else:
                            stat = 6
                else:
                    stat = 4
        elif "altcode" in request.POST:
            text = request.POST.get("altqr")
            if text:
                teams = Team.objects.all()
                for team in teams:
                    if team.Refral_Code == text:
                        if team.ispresent == False:
                            selected = team
                            team.ispresent = True
                            team.save()
                            stat = 1
                            break
                        else:
                            stat = 3
                    else:
                        stat = 8
            else:
                stat = 2
    return render(request, 'events/qr-code-verify.html', {'stats': stat, 'team': selected})

    # print(request.FILES['prof_img']


@event_head_present
def event_present(request):
    present = []
    notcame = []
    user = request.user
    if RoleAssignment.objects.filter(user=user).count:
        event_headpresent = RoleAssignment.objects.get(user=user).event
        event = EventMaster.objects.get(event_name=event_headpresent)
        teams = Team.objects.all()
        for team in teams:
            if team.receipt.event == event:
                if team.ispresent:
                    present.append(team)
                else:
                    notcame.append(team)
    if request.method == "POST":
        if "send" in request.POST:
            participants_selecteds = request.POST.getlist('participants')
            subject = request.POST.get('subject')
            message = request.POST.get('message')
            if len(request.FILES) == 1:
                myfile = request.FILES['file']
                fs = FileSystemStorage()
                filename = fs.save(myfile.name, myfile)
                uploaded_file_url = fs.path(filename)
                for participants_selected in participants_selecteds:
                    send_email(participants_selected, subject, message, [uploaded_file_url])
            else:
                for participants_selected in participants_selecteds:
                    send_email(participants_selected, subject, message)
            participants1 = Team.objects.values_list('user', flat=True).distinct().all()
            participants = []
            for p in participants1:
                participants.append(MyUser.objects.get(pk=p))
        elif "resend" in request.POST:
            participants_selected = request.POST.getlist('participants')
            for obj in participants_selected:
                print(obj)
                selected_user = MyUser.objects.get(email=obj)
                team_selecteds = Team.objects.filter(user=selected_user)
                for team_selected in team_selecteds:
                    if team_selected.receipt.event == event:
                        transaction = Transaction.objects.get(receipt=team_selected.receipt)
                    else:
                        transaction = None
                mail_subject = 'You have registered for ' + event.event_name + ' '
                message = render_to_string('events/receiptCashPayment.html', {
                    'user': selected_user,
                    'event': event,
                    'team': team_selected,
                    'transaction': transaction,
                })

                send_email(selected_user.email, mail_subject, message, [team_selected.QRcode.path])

    return render(request, "events/event--presenty.html", {'presents': present, 'notcomes': notcame})


@require_GET
def web_push(request):
    webpush_settings = getattr(settings, 'WEBPUSH_SETTINGS', {})
    vapid_key = webpush_settings.get('VAPID_PUBLIC_KEY')
    user = request.user
    return render(request, 'user/web_push.html', {user: user, 'vapid_key': vapid_key})


@require_POST
@csrf_exempt
def send_push(request):
    try:
        body = request.body
        data = json.loads(body)

        if 'head' not in data or 'body' not in data or 'id' not in data:
            return JsonResponse(status=400, data={"message": "Invalid data format"})

        user_id = data['id']
        user = get_object_or_404(User, pk=user_id)
        payload = {'head': data['head'], 'body': data['body']}
        send_user_notification(user=user, payload=payload, ttl=1000)

        return JsonResponse(status=200, data={"message": "Web push successful"})
    except TypeError:
        return JsonResponse(status=500, data={"message": "An error occurred"})


# @event_head_present
def event_count(request):
    role = RoleAssignment.objects.get(user=request.user)
    count = Receipt.objects.filter(event=role.event).count() - 1
    return render(request, 'user/event-count.html', {'count': count, 'event': role.event})


def jsonview(request):
    # str = {"employees":[{"firstname":"John","age":30,"mail":"john@gmail.com"},{"firstname":"Jimmy","age":25,"mail":"jimmy@gmail.com"},{"firstname":"Jenny","age":22,"mail":"jenny@gmail.com"},{"firstname":"Jeremy","age":40,"mail":"jeremy@gmail.com"},{"firstname":"Justin","age":32,"mail":"justin@gmail.com"}]}
    # todis = json.dumps(str)
    return render(request, 'events/package.json', {})

def verify_qr_feedback(request):
    if request.method == "POST":
        text = request.POST.get("altqr")
        check = 0
        if text:
            teams = Team.objects.all()
            curteam = None
            for team in teams:
                if team.Refral_Code == text:
                   curteam = team
                   check = 1
                   break
        if check == 1:
            questions = Feedback_questions.objects.all()
            options = Feedback_options.objects.all()
            option = Feedback_options.objects.latest('pk')
            return render(request, 'events/feedback.html', {'team': curteam , 'questions': questions , 'option' : option})
        else:
            return render(request, 'events/verify-feedback.html',{'error': "No entry Found"})
    else :
        return render(request, 'events/verify-feedback.html')

def feedback(request):
    if request.method == "POST":
       questions = Feedback_questions.objects.all()
       team = Team.objects.get(pk=request.POST.get('team'))
       for question in questions:
           print(request.POST.get('radio'+str(question.pk)))
           feed = Feedback()
           feed.team = team
           feed.question = question
           feed.option = Feedback_options.objects.get(pk = request.POST.get('radio'+str(question.pk)))
           feed.save()
       feed_comment = Feedback_comments()
       feed_comment.team = team
       if request.POST.get('comment'):
        feed_comment.comment = request.POST.get('comment')
       if request.POST.get('name1'):
        feed_comment.name1 = request.POST.get('name1')
       if request.POST.get('number1'):
        feed_comment.number1 = request.POST.get('number1')
       if request.POST.get('name2'):
        feed_comment.name2 = request.POST.get('name2')
       if request.POST.get('number2'):
        feed_comment.number2 = request.POST.get('number2')
       if request.POST.get('name3'):
        feed_comment.name3 = request.POST.get('name3')
       if request.POST.get('number3'):
        feed_comment.number3 = request.POST.get('number3')
       if request.POST.get('name4'):
         feed_comment.name4 = request.POST.get('name4')
       if request.POST.get('number4'):
        feed_comment.number4 = request.POST.get('number4')
       feed_comment.save()
       return render(request, 'events/verify-feedback.html',{'done':1})
    else:
        questions = Feedback_questions.objects.all()
        option = Feedback_options.objects.latest('pk')
        return render(request, 'events/feedback.html', { 'questions': questions, 'option': option})